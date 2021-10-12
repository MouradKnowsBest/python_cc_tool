import shutil
import pandas as pd
from tkinter import *
from tkinter import filedialog, ttk
from os import walk

def get_data_from_excel():
    global ce_que_on_veut_copier
    global path_data

    path_data = filedialog.askopenfilename(initialdir="C://Users//p094836//Desktop//", title="choose your Excell file",
                                            filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))

    df = pd.read_excel(path_data, sheet_name=0)  # can also index sheet by name or fetch all sheets
    ce_que_on_veut_copier = df["data"].tolist()
    print("Nombre de fichier chargés : {}".format(len(ce_que_on_veut_copier)))
    pathlabel_1.config(text="Nombre de fichiers chargés : {}".format(len(ce_que_on_veut_copier)), fg="black", width=45)
    pathlabel_1.place(relx=.10, rely=.27)

    return ce_que_on_veut_copier

def choose_extension(*args):
    global choosed_extension
    choosed_extension = variable_extension.get()
    print("variable changed!, Choosed extension : {}".format(choosed_extension))
    return choosed_extension

def get_chemin_source():
    global path_source, les_fichiers_source
    path_source = filedialog.askdirectory()
    print("Source path : {}".format(path_source))

    pathlabel_3.config(text="Chemin Source : {}".format(path_source), fg="red")  # , width=80)
    pathlabel_3.place(relx=.05, rely=.62)

    if variable_extension.get() == ".csv":
            dir = []
            les_dossiers = []
            les_fichiers_source = []
            for (et, er, filenames) in walk(path_source):
                dir.extend(et)
                les_dossiers.extend(er)
                les_fichiers_source.extend(filenames)
                break

    elif variable_extension.get() == ".mat":
        dir = []
        les_dossiers = []
        les_fichiers_source = []
        for (et, er, filenames) in walk(path_source):
            dir.extend(et)
            les_dossiers.extend(er)
            les_fichiers_source.extend(filenames)
            break

    elif variable_extension.get() == ".emp4":
        dir = []
        les_dossiers = []
        les_fichiers_source = []
        for (et, er, filenames) in walk(path_source):
            dir.extend(et)
            les_dossiers.extend(er)
            les_fichiers_source.extend(filenames)
            break

    elif variable_extension.get() == "":
        dir = []
        les_dossiers = []
        les_fichiers_source = []
        for (et, er, filenames) in walk(path_source):
            dir.extend(et)
            les_dossiers.extend(er)
            les_fichiers_source.extend(filenames)
            break

    else:
        print("No extention given")

def get_chemin_destination():
    global path_destination, les_fichiers_destination
    path_destination = filedialog.askdirectory()
    print("Destination path : {}".format(path_destination))

    pathlabel_4.config(text="Chemin Destination : {}".format(path_destination), fg="green")# , width=80)
    pathlabel_4.place(relx=.05, rely=.66)

    if variable_extension.get() == ".csv":
            dir = []
            les_dossiers = []
            les_fichiers_destination = []
            for (et, er, filenames) in walk(path_source):
                dir.extend(et)
                les_dossiers.extend(er)
                les_fichiers_destination.extend(filenames)
                break

    elif variable_extension.get() == ".mat":
        dir = []
        les_dossiers = []
        les_fichiers_destination = []
        for (et, er, filenames) in walk(path_source):
            dir.extend(et)
            les_dossiers.extend(er)
            les_fichiers_destination.extend(filenames)
            break

    elif variable_extension.get() == ".emp4":
        dir = []
        les_dossiers = []
        les_fichiers_source = []
        for (et, er, filenames) in walk(path_source):
            dir.extend(et)
            les_dossiers.extend(er)
            les_fichiers_source.extend(filenames)
            break

    else:
        print("No extention given")
    return path_destination

def get_text():
    global texte_a_rajouter, choosed_extension
    texte_a_rajouter = ent.get()
    choosed_extension = variable_extension.get()
    print("texte_a_rajouter : {}".format(texte_a_rajouter))
    print("variable changed!, Choosed extension : {}".format(choosed_extension))

def couper_fichiers():
    """
    :param original_path: Path of the original data from which we are going to take data
    :param destination_path: Path to the destiniation where we are going to put the data
    :param ce_que_on_veut_copier: A list containing the data
    :param extension: The extension of the data
    :return: The number of data correctly/incorrectly processed
    """

    #  The case of Cutting

    if variable_couper.get():
        extension = variable_extension.get()
        print(extension)
        print("Le texte à rajouter : {}".format(texte_a_rajouter))
        #extension = ".csv"
        counto = 0
        for z in list(ce_que_on_veut_copier):
                try:
                    shutil.move(path_source + "//" + z + texte_a_rajouter + extension,
                                path_destination + "//" + z + texte_a_rajouter + extension)
                    print("{}//{}{}{}".format(path_source, z, texte_a_rajouter, extension))
                    print("Fichier non trouvé, on le saute et on continue...")
                    counto += 1
                    pathlabel_2.config(text="Avancement du transfert des données : {}".format(str(counto)), fg="black", width=45)
                    pathlabel_2.place(relx=.1, rely=.80)
                    pathlabel_2.update()  # push the change to the screen
                except FileNotFoundError:
                    print("{}//{}{}".format(path_source, z, extension))
                    print("Fichier non trouvé, on le saute et on continue...")
                    pathlabel_5.config(text="Fichier(s) non trouvé(s) !", fg="red", width=45)
                    pathlabel_5.place(relx=.250, rely=.785)
                    pathlabel_5.update()  # push the change to the screen
                    continue
                except TypeError:
                        #print(original_path + z + extension)
                        print("Type ERROR, on le saute et on continue...")
                        continue

    #  The case of Copy

    elif variable_copier.get():
        extension = variable_extension.get()
        print(extension)
        print("Le texte à rajouter : {}".format(texte_a_rajouter))
        #extension = ".csv"

        counto = 0
        for z in list(ce_que_on_veut_copier):
                try:
                    shutil.copy(path_source + "//" + z + texte_a_rajouter + extension,
                                path_destination + "//" + z + texte_a_rajouter + extension)
                    counto += 1
                    pathlabel_2.config(text="Avancement du renommage : " + str(counto), fg="black", width=45)
                    pathlabel_2.place(relx=.12, rely=.80)
                    pathlabel_2.update()  # push the change to the screen
                except FileNotFoundError:
                    print(path_source + "//" + z + texte_a_rajouter + extension)
                    print("Fichier non trouvé, on le saute et on continue...")
                    continue
                except TypeError:
                        #print(original_path + z + extension)
                        print("Type ERROR, on le saute et on continue...")
                        continue

    #  No Option has been chosen

    else:
        print("Aucune option (copie ou couper) choisie !")

    return texte_a_rajouter

def compare():
    import numpy as np
    ce_que_on_veut_copier_ext = [i + str(texte_a_rajouter) + str(choosed_extension) for i in ce_que_on_veut_copier]
    ce_qui = np.setdiff1d(ce_que_on_veut_copier_ext, les_fichiers_source)
    ce_qui_safia = [i[:-4] for i in ce_qui]

    print("Ce qui reste : {}".format(ce_qui_safia))

    from openpyxl import load_workbook, Workbook

    wb = Workbook()
    sheets = wb.sheetnames
    ws1 = wb[sheets[0]]
    #ws1 = wb.get_sheet_by_name('Sheet1')
    ws1 = wb.active
    ws1['A1'] = 'data'

    for r in range(0, len(ce_qui_safia)):
        ws1.cell(row=r + 2, column=1).value = ce_qui_safia[r]

    wb.save("{}//Ce_qui_reste.xlsx".format(path_source))
    #compare(ce_que_on_veut_copier, les_fichiers_source)


root = Tk()
root.title('Outil pour Copier/Couper des fichiers - AKKA Technologies V5')
root.geometry("600x435")
root.resizable(width=False, height=False)

pathlabel_1 = Label(root)
pathlabel_2 = Label(root)
pathlabel_3 = Label(root)
pathlabel_4 = Label(root)
pathlabel_5 = Label(root)


d = Button(root, text='Charger les données', borderwidth=1, relief=SOLID, command=lambda: get_data_from_excel())
d.place(x=130, y=90)
d.config(width=17)

variable_extension = StringVar()
numberChosen = ttk.Combobox(root, width=8, textvariable=variable_extension)
numberChosen['values'] = (".mat", ".csv", ".emp4", "")
numberChosen.place(x=275, y=91.30)
numberChosen.current(0)

ent = Entry(root)
ent.place(x=355, y=92)
fok = Button(root, text='OK', borderwidth=1, relief=SOLID, command=lambda: get_text())
fok.place(x=490, y=82)
fok.config(width=5, height=2)


rt = Label(root, text="Chargez les données, choisissez l'extention, puis cliquez sur OK")
rt.place(x=150, y=45)

rt = Label(root, text="→", font=("Helvetica", 40))
rt.place(x=265, y=195)

e = Button(root, text='Chemin source', bg="lightsalmon", borderwidth=1, relief=SOLID, command=lambda: get_chemin_source())
e.place(x=45, y=162)
e.config(width=17, height=3)

f = Button(root, text='Chemin destination', bg="mediumspringgreen", borderwidth=1, relief=SOLID, command=lambda: get_chemin_destination())
f.place(x=425, y=162)
f.config(width=17, height=3)

f = Button(root, text='Comparaison', bg="snow", borderwidth=1, relief=SOLID, command=lambda: compare())
f.place(x=242, y=170)
f.config(width=15, height=2)

variable_copier = IntVar()
a = Checkbutton(root, text="copier", variable=variable_copier)
a.place(x=220, y=320)

variable_couper = IntVar()
b = Checkbutton(root, text="couper", variable=variable_couper)
b.place(x=310, y=320)


c = Button(root, text="Transférer les données", bg='lightsteelblue', borderwidth=1, relief=SOLID, command=lambda: couper_fichiers())
c.place(x=160, y=370)
c.config(height=2, width=37)

root.mainloop()